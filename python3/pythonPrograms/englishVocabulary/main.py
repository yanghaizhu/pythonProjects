# Information: help you to remamber IELTS vocabulary.
#
# 1. IELTS vocabulary saved in vocabulary.csv;
# 2. Save the vocabulary.csv as vocabulary.xlsx;
# 3. Use the pygame module to control show the vocabulary one by one;
# 4. Use openpyxl module to read from vocabulary, and record we remembered word.
# 5. Assume the one you skip without looking for detail information is the
#    one you have remembered.
# 6. LeftClick Mouce will show more information, rightClick mouce will
#    skip to next no-remembered word.
# 7. Use speech module to read the word and the sentence
# 8. Use threading module to control click while reading is not finished.

import sqlite3
from datetime import datetime

import pygame
from pygame.locals import *
from pygame.color import THECOLORS
from openpyxl import *
from openpyxl.styles import *
from collections import *
import speech
import threading
#from os import *
import os
from enum import Enum
import warnings
warnings.filterwarnings('ignore')

def readOutLoudly(sentence:str):
    speech.say(sentence)


class columEnum(Enum):
    Vocabulary = 1
    Pronounce = 2
    Chinese = 3
    Split = 4
    RememberTips = 5
    RememberImage = 6
    Sentence = 7
    SentenceChinese = 8
    RecordTimes = 9
    RecordDone = 10

class VOCABULARY:
    global bookmarkRecordPosRow
    global bookmarkRecordPosCol
    bookmarkRecordPosRow = 1
    bookmarkRecordPosCol= 15
    def __init__(self,fileName:str,sheetName:str):
        self.fileName = fileName
        self.sheetName = sheetName
        self.wb = load_workbook(fileName)
        self.ws = self.wb.get_sheet_by_name(sheetName)
        #how if no value, or not int type
        self.iRow = self.ws.cell(bookmarkRecordPosRow,bookmarkRecordPosCol).value
        self.row = self.ws[self.iRow]

    def activeAnotherSheet(self,sheetName:str):
        self.sheetName = sheetName
        self.wb = load_workbook(self.fileName)
        self.ws = self.wb.get_sheet_by_name(sheetName)
        self.iRow = 2
        self.row = self.ws[self.iRow]
        return self.row

    @property
    def indexRow(self,):
        return self.iRow
    @property
    def nextRow(self,):
        # read the first vocabulary from table.
        iRow = self.iRow
        iRow = iRow + 1
        while True:
            row = self.ws[iRow]
            # If we had remembered this word, read next
            if row[columEnum.RecordDone.value].value == 'YES':
                iRow = iRow + 1
                continue
            self.iRow = iRow
            self.row = row
            myDict = dict()
            myDict["RecordTimes"] = row[columEnum.RecordTimes.value].value
            myDict["Vocabulary"] = row[columEnum.Vocabulary.value].value
            myDict["Pronounce"] = row[columEnum.Pronounce.value].value
            myDict["Split"] = row[columEnum.Split.value].value
            myDict["Chinese"] = row[columEnum.Chinese.value].value
            myDict["RememberTips"] = row[columEnum.RememberTips.value].value
            myDict["RememberImage"] = row[columEnum.RememberImage.value].value
            myDict["Sentence"] = row[columEnum.Sentence.value].value
            myDict["SentenceChinese"] = row[columEnum.SentenceChinese.value].value
            return myDict

    def saveRemember(self,doneFlag,recordTimes):
        row = self.row
        row[columEnum.RecordDone.value].value = doneFlag
        row[columEnum.RecordTimes.value].value = recordTimes
        self.wb.save(self.fileName)

    def updateBookmark(self,):
        self.ws.cell(bookmarkRecordPosRow,bookmarkRecordPosCol).value = self.iRow
        self.wb.save(self.fileName)

class PYGAME_FRAME:
    def __init__(self, pygame, screenSize, title):
        pygame.init()
        pygame.mixer.init()
        pygame.display.set_caption(title)
        self.screenSize = screenSize
        self.pygame = pygame
        self.screen = pygame.display.set_mode(screenSize, pygame.DOUBLEBUF, 32)
        self.fontDict = dict()
        self.imgDict = dict()
        self.fontDict["default"] = self.pygame.font.Font('fonts/arial.ttf', 10)

    def addFond(self,name,fontPath,size):
        self.fontDict[name] = self.pygame.font.Font(fontPath, size)

    def addImag(self,name,imgPath,aphal):
        self.imgDict[name] = self.pygame.image.load(imgPath)
        if (aphal < 100):
            self.imgDict[name].set_alpha(aphal)


    def resizeImag(self,name,size):
        print(name)
        print(self.imgDict)
        if(name in self.imgDict.keys()):
            self.imgDict[name] = self.pygame.transform.scale(self.imgDict[name], size)
    def blitText(self,fontName,textStr,color,position,centFlag):
        text = self.fontDict[fontName].render(textStr, True, color)
        w, h = text.get_size()
        if (centFlag):
            posStart = (position[0]-w/2, position[1]-h/2)
        else:
            posStart = position
        self.screen.blit(text, posStart)
    def updateDisp(self,):
        self.pygame.display.update()
    def fillbgColor(self,color):
        self.screen.fill(color)
    def blitImg(self,name,pos):
        self.screen.blit(self.imgDict[name], pos)
class vocabularyDict:
    def __init__(self,app,vocabularyDict):
        self.app = app
        self.vocab = vocabularyDict
    def showVocab(self,color,pos_x,clickCnt):
        self.app.blitText("font1", self.vocab["Vocabulary"], color, (pos_x, 40), True)
        self.app.blitText("font2", self.vocab["Pronounce"], color, (pos_x, 80), True)
        if(clickCnt > 0):
            app.blitText("font3", self.vocab["Chinese"], color, (pos_x, 120), True)
        if(clickCnt > 1):
            app.blitText("font3", self.vocab["Split"], color, (pos_x, 160), True)
        if(clickCnt > 2):
            app.blitText("font3", self.vocab["RememberImage"], color, (pos_x, 190), True)
        if(clickCnt > 3):
            app.blitText("font3", self.vocab["Sentence"], color, (pos_x, 230), True)
        if(clickCnt > 4):
            app.blitText("font3", self.vocab["SentenceChinese"], color, (pos_x, 260), True)

    def showRecordInfo(self,infoStr,color,pos):
        app.blitText("font3", infoStr, color,pos, False)


if __name__ == '__main__':

    conn = sqlite3.connect('vocabulary.db')
    c = conn.cursor()
#    c.execute('''DROP TABLE IF EXISTS vocabulary''')
    # ID|时间|类型|名称|作者|排行|图片
#    c.execute('''CREATE TABLE IF NOT EXISTS vocabulary (ID INTEGER PRIMARY KEY, Vocabulary TEXT, Pronounce TEXT, Split TEXT, Chinese TEXT, RememberTips TEXT, RememberImage TEXT, Sentence TEXT, SentenceChinese TEXT)''')

    # vocabulary table related parameters
    leftClickNumber = 0
    recordTimes = 0
    recordDone = 0
    rowNumber = 5473
    WINDOW_W, WINDOW_H = 800, 500
    blackColor = (0,0,0)
    whiteColor = (255,255,255)
    screenSize = (WINDOW_W, WINDOW_H)

    quitFlag = True
    leftClickCnt = 0

    # open vocabulary.csv, and save as vocabulary.xlsx, then it can work normally.
    myVocabulary = VOCABULARY('vocabulary.xlsx','vocabulary')
    myDict = myVocabulary.nextRow


    app = PYGAME_FRAME(pygame,screenSize,"背单词")
    app.addImag("BG","bg.jpg",40)
    app.resizeImag("BG",(800,1500))
    app.addFond("font1",'fonts/msyh.ttc', 35)
    app.addFond("font2",'fonts/arial.ttf', 30)
    app.addFond("font3",'fonts/msyh.ttc', 15)

    app.fillbgColor(blackColor)
    app.blitImg("BG",(0,0))

    vocabShowInApp = vocabularyDict(app,myDict)
    vocabShowInApp.showVocab(whiteColor,WINDOW_W/2, leftClickCnt)
    vocabShowInApp.showRecordInfo("正在学习第" + str(myVocabulary.indexRow) + "/" + str(rowNumber) + "个单词...",whiteColor,(0,0))

    app.updateDisp()

    t1 = threading.Thread(target=readOutLoudly, args=(myDict["Vocabulary"],))
    t1.start()

    a = 1
    while quitFlag:
        a = a + 1
        ## change to db start
#        myDict = myVocabulary.nextRow
#        Pronounce = myDict["Pronounce"].replace("\'", "\'\'").strip('\n').strip('\r')
#        Sentence = myDict["Sentence"].replace("\'", "\'\'").strip('\n').strip('\r')
#        SentenceChinese = myDict["SentenceChinese"].replace("\'", "\'\'").strip('\n').strip('\r')
#        Vocabulary = myDict["Vocabulary"].replace("\'", "\'\'").strip('\n').strip('\r')
#        Split = myDict["Split"].replace("\'", "\'\'").strip('\n').strip('\r')
#        Chinese = myDict["Chinese"].replace("\'", "\'\'").strip('\n').strip('\r')
#        RememberTips = myDict["RememberTips"].replace("\'", "\'\'").strip('\n').strip('\r')
#        RememberImage = myDict["RememberImage"].replace("\'", "\'\'").strip('\n').strip('\r')
#        sql = "INSERT INTO vocabulary(Vocabulary,Pronounce,Split,Chinese,RememberTips,RememberImage,Sentence,SentenceChinese) VALUES " \
#              "('" + Vocabulary + "','" + Pronounce + "','" + Split + "','" + Chinese\
#              + "','" + RememberTips + "','" + RememberImage + "','" + Sentence + "','" + SentenceChinese + "')"
#        print(sql)
#        c.execute(sql)
        #change to db end
#        if(a >= 5472):
#            sql = "SELECT * FROM vocabulary"
#            c.execute(sql)
#            retrievedData = c.fetchall()
#            print(retrievedData[0][0])
#            # Save (commit) the changes
#            conn.commit()
            # Close the connection
#            conn.close()
#            quitFlag = False
#        continue

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                quitFlag = False
            elif event.type == MOUSEMOTION:
                #return the X and Y position of the mouse cursor
                pos = pygame.mouse.get_pos()
                mouse_x = pos[0]
                mouse_y = pos[1]
            elif event.type == MOUSEBUTTONDOWN:
                pressed_array = pygame.mouse.get_pressed()
                for index in range(len(pressed_array)):
                    if pressed_array[index]:
                        if index == 0:
                            app.fillbgColor(blackColor)
                            app.blitImg("BG",(0,0))
                            #print('Pressed LEFT Button!')
                            leftClickCnt = leftClickCnt + 1
                            vocabShowInApp = vocabularyDict(app, myDict)
                            vocabShowInApp.showVocab(whiteColor, WINDOW_W / 2, leftClickCnt)
                            vocabShowInApp.showRecordInfo("正在学习第" + str(myVocabulary.indexRow) + "/" + str(rowNumber) + "个单词...",whiteColor,(0,0))

                            t1.join()
                            app.updateDisp()

                            if leftClickCnt > 3:
                                t1 = threading.Thread(target=readOutLoudly, args=(myDict["Sentence"],))
                                t1.start()
                            else:
                                t1 = threading.Thread(target=readOutLoudly, args=(myDict["Vocabulary"],))
                                t1.start()
                        elif index == 1:
                            t1.join()
                            #print('The mouse wheel Pressed!')
                        elif index == 2:
                            myDict = myVocabulary.nextRow
                            t1.join()
                            #print('Pressed RIGHT Button!')
                            leftClickCnt = 0
                            app.fillbgColor(blackColor)
                            app.blitImg("BG",(0,0))
                            vocabShowInApp = vocabularyDict(app, myDict)
                            vocabShowInApp.showVocab(whiteColor, WINDOW_W / 2, leftClickCnt)
                            vocabShowInApp.showRecordInfo("正在学习第" + str(myVocabulary.indexRow) + "/" + str(rowNumber) + "个单词...",whiteColor,(0,0))

                            app.updateDisp()

                            if leftClickCnt < 4:
                                myVocabulary.saveRemember("YES",recordTimes+1)

                            t1 = threading.Thread(target=readOutLoudly, args=(myDict["Vocabulary"],))
                            t1.start()
    myVocabulary.updateBookmark()# where to start first

    exit()